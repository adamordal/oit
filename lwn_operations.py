import csv
from openpyxl import Workbook

def read_csv(file_path):
    agency_dict = {'doc': [], 'cdec': [], 'hc': [], 'vmc': []}
    with open(file_path, mode='r') as file:
        reader = csv.reader(file)
        for _ in range(4):
            next(reader)  # Skip the first 4 rows

        headers = next(reader)  # The 5th row is the header
        agency_index = headers.index('agencyName')

        for row in reader:
            if len(row) > agency_index:  # Ensure the row has enough columns
                agency_name = row[agency_index].strip().lower()  # Make agency name case insensitive
                row_dict = {headers[i]: row[i] for i in range(len(headers))}
                if agency_name not in agency_dict:
                    agency_dict[agency_name] = []
                agency_dict[agency_name].append(row_dict)
    return agency_dict

def reassign_other_agency_rows(agency_dict):
    other_agency_rows = agency_dict.get('other', [])
    for key in agency_dict.keys():
        if key != 'other' and key != 'vmc':
            matches_to_remove = []
            for match in other_agency_rows:
                if key.lower() in match['Storage Policy'].lower():
                    agency_dict[key].append(match)
                    matches_to_remove.append(match)
            for match in matches_to_remove:
                other_agency_rows.remove(match)
                    
    return agency_dict

def assign_vmc_agency_rows(agency_dict):
    vmc_agency_rows = []
    for key, rows in agency_dict.items():
        if key == 'vmc':
            continue
        for row in rows:
            if 'vmc' in row.get('Storage Policy', '').lower():
                vmc_agency_rows.append(row)
    agency_dict['vmc'].extend(vmc_agency_rows)
    agency_dict_with_vmc = agency_dict.copy()
    return agency_dict_with_vmc

def remove_test_entries_from_gov(agency_dict):
    if 'gov' not in agency_dict:
        return agency_dict

    gov_agency_rows = agency_dict['gov']
    agency_dict['gov'] = [row for row in gov_agency_rows if 'test' not in row.get('Storage Policy', '').lower()]
    return agency_dict

def write_agency_to_csv(agency_rows, output_file_path):
    if not agency_rows:
        return

    headers = agency_rows[0].keys()
    with open(output_file_path, mode='w', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writeheader()
        for row in agency_rows:
            writer.writerow(row)

def write_agencies_to_excel(agency_dict, output_file_path):
    workbook = Workbook()
    for agency, rows in agency_dict.items():
        if not rows:
            continue
        sheet = workbook.create_sheet(title=agency)
        headers = list(rows[0].keys())  # Convert dict_keys to list
        sheet.append(headers)
        for row in rows:
            sheet.append([row[header] for header in headers])
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']  # Remove the default sheet
    workbook.save(output_file_path)

       

def calc_department_data(agency_dict):
    calculated_data = {}
    for department, rows in agency_dict.items():
        if department == 'vmc':  # Exclude 'vmc' department
            continue

        primary_gb = 0
        primary_scale_gb = 0
        lw_primary_gb = 0
        lw_cloud_gb = 0
        cloud_gb = 0
        cloud_selective_gb = 0
        ef_cloud_gb = 0

        for row in rows:
            if row['Copy'] == 'Primary':
                primary_gb += float(row['All Data on Media'])
            elif 'pri' in row['Copy'].lower() and 'scaleprotect' in row['Copy'].lower():
                primary_scale_gb += float(row['All Data on Media'])
            elif row['Copy'] == 'LW Primary':
                lw_primary_gb += float(row['All Data on Media'])
            elif 'lw cloud' in row['Copy'].lower():
                lw_cloud_gb += float(row['All Data on Media'])
            elif row['Copy'] == 'Cloud':
                cloud_gb += float(row['All Data on Media'])
            elif row['Copy'] == 'Cloud Selective':
                cloud_selective_gb += float(row['All Data on Media'])
            elif 'ef cloud' in row['Copy'].lower():
                ef_cloud_gb += float(row['All Data on Media'])

        primary_total = primary_gb + primary_scale_gb + lw_primary_gb
        cloud_total = cloud_gb + cloud_selective_gb + lw_cloud_gb + ef_cloud_gb

        calculated_data[department] = {
            'primary_gb': primary_gb,
            'primary_scale_gb': primary_scale_gb,
            'lw_primary_gb': lw_primary_gb,
            'lw_cloud_gb': lw_cloud_gb,
            'cloud_gb': cloud_gb,
            'cloud_selective_gb': cloud_selective_gb,
            'ef_cloud_gb': ef_cloud_gb,
            'primary_total': primary_total,
            'cloud_total': cloud_total
        }

    return calculated_data
