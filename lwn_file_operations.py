import csv
from openpyxl.styles import Border, Side, Alignment, Font
import openpyxl

def csv_to_dict(file_path):
    with open(file_path, mode='r') as file:
        # Skip the first 4 rows
        for _ in range(4):
            next(file)
        reader = csv.DictReader(file)
        # Read the rest of the file into a dictionary
        data = [row for row in reader]
    return data

def sort_data(criteria, data):
    # Create a copy of the list
    data_copy = data[:]
    
    # Sort the data based on the criteria
    sorted_data = sorted(data_copy, key=lambda x: criteria.index(x['agencyName']) if x['agencyName'] in criteria else len(criteria))
    
    # Group sorted data by 'agencyName'
    grouped_data = {}
    for item in sorted_data:
        agency = item['agencyName']
        if agency not in grouped_data:
            grouped_data[agency] = []
        grouped_data[agency].append(item)
    
    return grouped_data

def calc_department_data(department, sorted_data):
    primary_gb = 0
    primary_scale_gb = 0
    lw_primary_gb = 0
    lw_cloud_gb = 0
    cloud_gb = 0
    cloud_selective_gb = 0
    ef_cloud_gb = 0

    for row in sorted_data[department]:
        if row['Copy'] == 'Primary':
            primary_gb += float(row['All Data on Media'])
        elif 'pri' in row['Copy'].lower() and 'scaleprotect' in row['Copy'].lower():
            primary_scale_gb += float(row['All Data on Media'])
        elif row['Copy'] == 'LW Primary':
            lw_primary_gb += float(row['All Data on Media'])
        elif 'lw cloud' in row['Copy'].lower():
            lw_cloud_gb += float(row['All Data on Media'])
        elif row['Copy'] == 'Cloud':
            cloud_gb += float(row['All Data on Media'])
        elif row['Copy'] == 'Cloud Selective':
            cloud_selective_gb += float(row['All Data on Media'])
        elif 'ef cloud' in row['Copy'].lower():
            ef_cloud_gb += float(row['All Data on Media'])

    return {
        'primary_gb': primary_gb,
        'primary_scale_gb': primary_scale_gb,
        'lw_primary_gb': lw_primary_gb,
        'lw_cloud_gb': lw_cloud_gb,
        'cloud_gb': cloud_gb,
        'cloud_selective_gb': cloud_selective_gb,
        'ef_cloud_gb': ef_cloud_gb
    }

def create_sheet(wb, department, rows, medium_border):
    if department not in wb.sheetnames:
        ws = wb.create_sheet(title=department)
    else:
        ws = wb[department]

    start_row = 1
    start_col = 1  # Column A

    headers = rows[0].keys()
    for col, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.border = medium_border
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row_data in enumerate(rows, start=start_row + 1):
        for col_idx, (key, value) in enumerate(row_data.items(), start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = medium_border
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    # Apply medium border to the outside of the table
    for col in range(start_col, start_col + len(headers)):
        ws.cell(row=start_row, column=col).border = medium_border
        ws.cell(row=start_row + len(rows), column=col).border = medium_border
    for row in range(start_row, start_row + len(rows) + 1):
        ws.cell(row=row, column=start_col).border = medium_border
        ws.cell(row=row, column=start_col + len(headers) - 1).border = medium_border

def write_chargeback_data(chargeback_ws, sorted_data, medium_border):
    start_row = 32
    start_col = 2  # Column B

    for i, (department, rows) in enumerate(sorted_data.items()):
        if department == 'vmc':
            continue  # Skip writing vmc data
        row = start_row + i
        department_data = calc_department_data(department, sorted_data)
        primary_total = department_data['primary_gb'] + department_data['primary_scale_gb'] + department_data['lw_primary_gb']
        cloud_total = department_data['cloud_gb'] + department_data['cloud_selective_gb'] + department_data['lw_cloud_gb'] + department_data['ef_cloud_gb']

        for col in range(start_col, start_col + 6):
            if col == start_col:
                value = department
            elif col == start_col + 1:
                value = primary_total
            elif col == start_col + 2:
                value = cloud_total
            elif col == start_col + 3:
                value = f'=C{row}+D{row}'
            elif col == start_col + 4:
                value = f'=C{row}/C{start_row + len(sorted_data) - 1}'
            else:
                value = f'=F{row}*$F$4'

            cell = chargeback_ws.cell(row=row, column=col, value=value)
            cell.border = medium_border
            if col == start_col:
                cell.font = Font(bold=True)
            if col in (start_col + 1, start_col + 2, start_col + 3):
                cell.number_format = '#,##0.00'
            elif col == start_col + 4:
                cell.number_format = '0.00%'
            elif col == start_col + 5:
                cell.number_format = '$#,##0.00'

    # Add total row
    total_row = start_row + len(sorted_data) - 1  # Adjust for skipped vmc
    chargeback_ws.cell(row=total_row, column=start_col, value="Total").border = medium_border
    chargeback_ws.cell(row=total_row, column=start_col).font = Font(bold=True)
    for col in range(start_col + 1, start_col + 4):
        cell = chargeback_ws.cell(row=total_row, column=col, value=f'=SUM({chr(64 + col)}{start_row}:{chr(64 + col)}{total_row - 1})')
        cell.border = medium_border
        cell.font = Font(bold=True)
        cell.number_format = '#,##0.00'

    # Apply medium border to the outside of the table
    for col in range(start_col, start_col + 6):
        chargeback_ws.cell(row=start_row, column=col).border = medium_border
        chargeback_ws.cell(row=total_row, column=col).border = medium_border
    for row in range(start_row, total_row + 1):
        chargeback_ws.cell(row=row, column=start_col).border = medium_border
        chargeback_ws.cell(row=row, column=start_col + 5).border = medium_border

def save_to_excel(sorted_data, template_path, output_path):
    wb = openpyxl.load_workbook(template_path)
    
    medium_border = Border(left=Side(style='medium'), 
                           right=Side(style='medium'), 
                           top=Side(style='medium'), 
                           bottom=Side(style='medium'),
                           outline=True)

    for department, rows in sorted_data.items():
        if department == 'vmc':
            continue  # Skip writing vmc data
        create_sheet(wb, department, rows, medium_border)

    # Write calculated department data to "Chargeback by Department" sheet
    chargeback_ws = wb["Chargeback By Department"]
    write_chargeback_data(chargeback_ws, sorted_data, medium_border)

    wb.save(output_path)

def main():
    # Example usage
    file_path = 'commvault.csv'
    template_path = 'template.xlsx'
    output_path = 'output.xlsx'

    criteria = [
        'vmc', 'cbms', 'cda', 'cdec', 'ecea', 'chs', 'hc', 'cdhs', 'cdle', 
        'cdor', 'cdot', 'cdphe', 'cst', 'cdps', 'dnr', 'dmva', 'doc', 'dola', 
        'dpa', 'dora', 'gov', 'hcpf', 'lottery', 'oit'
    ]
    #global data
    data = csv_to_dict(file_path)

    sorted_data = sort_data(criteria,data)
    
    save_to_excel(sorted_data, template_path, output_path)

if __name__ == "__main__":
    main()