import openpyxl
from openpyxl.styles import numbers
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import json
from extract_quota_usage import extract_quota_usage

def read_xlsx_to_dict(file_path, sheet_name, key_column, value_start_column):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]  # Access the sheet by name
    data = {}

    for row in sheet.iter_rows(min_row=2, max_row=101, values_only=True):
        key = row[key_column]
        if key is None:
            continue
        key = key.lower()
        values = row[value_start_column:]
        data[key] = values

    return data

def calculate_totals(departments, quota_usage_data):
    results = {}
    for department, sub_departments in departments.items():
        total_files = 0
        total_capacity = 0

        for sub_department in sub_departments:
            # Generate the criteria for the sub-department
            criteria = f"/ifs/{sub_department}".lower()

            # Calculate the total files and total physical capacity
            for quota in quota_usage_data:
                if quota["Path"].lower().startswith(criteria):
                    total_files += quota["Files"]
                    total_capacity += quota["Physical"]

        # Store the results in the dictionary
        results[department] = {
            'total_files': total_files,
            'total_capacity': total_capacity
        }
    return results

def write_to_template(data_dict_lwx, data_dict_lwn, LWX_Cost, LWN_Cost, output_path, departments_lwx, departments_lwn):
    # Load the template workbook
    template_path = os.path.join(os.path.dirname(__file__), 'template.xlsx')
    template_workbook = openpyxl.load_workbook(template_path)
    template_sheet = template_workbook.active

    # Define the cell locations for each department for LWX
    cell_locations_lwx = {
        'OIT': ('D8', 'E8'),
        'DPA': ('D9', 'E9'),   
        'CHS': ('D10', 'E10'),
        'DNR': ('D11', 'E11'),
        'SECOPS': ('D12', 'E12'),
        'DOLA': ('D13', 'E13'),
        'DORA': ('D14', 'E14'),
        'Public': ('D15', 'E15'),
        'DOR': ('D16', 'E16'),
        'CST': ('D17', 'E17'),
        'GOV': ('D18', 'E18'),
        'CDA': ('D19', 'E19'),
        'HCPF': ('D20', 'E20'),
        'CDOT': ('D21', 'E21'),
        'CDEC': ('D22', 'E22'),
        'CDPHE': ('D23', 'E23'),
        'CDLE': ('D24', 'E24'),
        'CDHS': ('D25', 'E25'),
        'Legislative': ('D26', 'E26')
    }

    # Define the cell locations for each department for LWN
    cell_locations_lwn = {
        'OIT': ('D32', 'E32'),
        'CDHS': ('D33', 'E33'),
        'DORA': ('D34', 'E34'),
        'CDOT': ('D35', 'E35')
    }

    # Write the usage data to the template for LWX
    for dept, (files_cell, usage_cell) in cell_locations_lwx.items():
        if dept not in data_dict_lwx:
            messagebox.showerror("Error", f"Department {dept} is missing in the LWX data source.")
            template_sheet[files_cell] = "N/A"
            template_sheet[usage_cell] = "N/A"
        else:
            total_files = data_dict_lwx[dept]['total_files']
            usage = data_dict_lwx[dept]['total_capacity']
            template_sheet[files_cell] = total_files
            template_sheet[usage_cell] = usage

    # Write the usage data to the template for LWN
    for dept, (files_cell, usage_cell) in cell_locations_lwn.items():
        if dept not in data_dict_lwn:
            messagebox.showerror("Error", f"Department {dept} is missing in the LWN data source.")
            template_sheet[files_cell] = "N/A"
            template_sheet[usage_cell] = "N/A"
        else:
            total_files = data_dict_lwn[dept]['total_files']
            usage = data_dict_lwn[dept]['total_capacity']
            template_sheet[files_cell] = total_files
            template_sheet[usage_cell] = usage

    # Add LWX_Cost and LWN_Cost to the template
    template_sheet['F3'] = float(LWX_Cost)
    template_sheet['F3'].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    template_sheet['F4'] = float(LWN_Cost)
    template_sheet['F4'].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # Remove external links
    if hasattr(template_workbook, 'remove_external_links'):
        template_workbook.remove_external_links()

    # Save the updated workbook
    template_workbook.save(output_path)
    print(f"Data written to {output_path}")

def browse_file(prompt, filetypes):
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(title=prompt, filetypes=filetypes)
    root.destroy()
    if not file_path:
        messagebox.showerror("Error", "No file selected!")
        raise ValueError("No file selected")
    return file_path

def save_file(prompt):
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.asksaveasfilename(title=prompt, defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    root.destroy()
    if not file_path:
        messagebox.showerror("Error", "No file selected!")
        raise ValueError("No file selected")
    return file_path

def main():
    try:
        # Define departments for LWX
        departments_lwx = {
            'OIT': ['OIT-LW', 'HQAdmins', 'DEPTS', 'OIT'],
            'SECOPS': ['SECOPS'],
            'DPA': ['DPA'],
            'CHS': ['CHS'],
            'DNR': ['DNR'],
            'DOLA': ['DOLA'],
            'DORA': ['DORA'],
            'Public': ['Public'],
            'DOR': ['DOR', 'Revenue'],
            'CST': ['CST'],
            'GOV': ['GOV'],
            'CDA': ['CDA'],
            'HCPF': ['HCPF'],
            'CDOT': ['CDOT', 'CDOTDMZ'],
            'CDEC': ['CDEC', 'CDECHIPAA'],
            'CDPHE': ['CDPHE'],
            'CDLE': ['CDLE'],
            'CDHS': ['CDHS', 'CDHSHIPAA'],
            'Legislative': ['Legislative']
        }

        # Define departments for LWN
        departments_lwn = {
            'OIT': ['OIT'],
            'CDHS': ['CDHS'],
            'DORA': ['DORA'],
            'CDOT': ['CDOT']
        }

        # Prompt the user to select the Ingram Micro cost report file
        ingram_file_path = browse_file("Select the Ingram Micro cost report file", [("Excel files", "*.xlsx")])

        # Load the Ingram Micro cost report
        sheet_name = 'Rating Report'
        data_dict = read_xlsx_to_dict(ingram_file_path, sheet_name, key_column=1, value_start_column=2)
        LWX_Cost = data_dict['lwx400'][20]
        LWN_Cost = data_dict['lwnl400'][20]

        # Prompt the user to select the LWX JSON configuration file
        json_file_path_lwx = browse_file("Select the LWX JSON configuration file", [("JSON files", "*.json")])

        # Load the LWX JSON configuration
        with open(json_file_path_lwx, 'r', encoding='utf8') as json_file:
            json_cfg_lwx = json.load(json_file)

        # Extract LWX quota usage data
        quota_usage_data_lwx = extract_quota_usage(json_cfg_lwx)

        # Prompt the user to select the LWN JSON configuration file
        json_file_path_lwn = browse_file("Select the LWN JSON configuration file", [("JSON files", "*.json")])

        # Load the LWN JSON configuration
        with open(json_file_path_lwn, 'r', encoding='utf8') as json_file:
            json_cfg_lwn = json.load(json_file)

        # Extract LWN quota usage data
        quota_usage_data_lwn = extract_quota_usage(json_cfg_lwn)

        # Calculate totals for LWX and LWN
        data_dict_lwx = calculate_totals(departments_lwx, quota_usage_data_lwx)
        data_dict_lwn = calculate_totals(departments_lwn, quota_usage_data_lwn)

        # Prompt the user to specify the output file name
        output_path = save_file("Save the output file as")

        # Write data to the template
        write_to_template(data_dict_lwx, data_dict_lwn, LWX_Cost, LWN_Cost, output_path, departments_lwx, departments_lwn)
    except ValueError as e:
        print(e)

if __name__ == "__main__":
    main()