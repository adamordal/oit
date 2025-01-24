import openpyxl
from openpyxl.styles import numbers, Border, Side
import os
from tkinter import filedialog, messagebox
import tkinter as tk


def read_xlsx_to_dict(file_path, sheet_name, key_column, value_start_column):
    # Read data from an Excel file and return it as a dictionary
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]  # Access the sheet by name
    data = {}

    for row in sheet.iter_rows(min_row=2, max_row=101, values_only=True):
        key = row[key_column]
        if key is None:
            continue
        key = key.lower()
        values = row[value_start_column:]
        data[key] = values

    return data

def browse_file(prompt, filetypes):
    # Open a file dialog to browse for a file
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(title=prompt, filetypes=filetypes)
    root.destroy()
    if not file_path:
        messagebox.showerror("Error", "No file selected!")
        raise ValueError("No file selected")
    return file_path

def save_file(prompt):
    # Open a file dialog to save a file
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.asksaveasfilename(title=prompt, defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    root.destroy()
    if not file_path:
        messagebox.showerror("Error", "No file selected!")
        raise ValueError("No file selected")
    return file_path

def write_to_template(data_dict_lwx, LWX_Cost, LWN_Cost, output_path, departments_lwx,lwn_final_data,calculated_data):
    # Load the template workbook
    template_path = os.path.join(os.path.dirname(__file__), 'template.xlsx')
    template_workbook = openpyxl.load_workbook(template_path)
    template_sheet = template_workbook.active

    # Define the cell locations for each department for LWX
    cell_locations_lwx = {
        'OIT': ('D8', 'E8'),
        'DPA': ('D9', 'E9'),   
        'CHS': ('D10', 'E10'),
        'DNR': ('D11', 'E11'),
        'SECOPS': ('D12', 'E12'),
        'DOLA': ('D13', 'E13'),
        'DORA': ('D14', 'E14'),
        'Public': ('D15', 'E15'),
        'DOR': ('D16', 'E16'),
        'CST': ('D17', 'E17'),
        'GOV': ('D18', 'E18'),
        'CDA': ('D19', 'E19'),
        'HCPF': ('D20', 'E20'),
        'CDOT': ('D21', 'E21'),
        'CDEC': ('D22', 'E22'),
        'CDPHE': ('D23', 'E23'),
        'CDLE': ('D24', 'E24'),
        'CDHS': ('D25', 'E25'),
        'Legislative': ('D26', 'E26')
    }


    # Write the usage data to the template for LWX
    for dept, (files_cell, usage_cell) in cell_locations_lwx.items():
        if dept not in data_dict_lwx:
            messagebox.showerror("Error", f"Department {dept} is missing in the LWX data source.")
            template_sheet[files_cell] = "N/A"
            template_sheet[usage_cell] = "N/A"
        else:
            total_files = data_dict_lwx[dept]['total_files']
            usage = data_dict_lwx[dept]['total_capacity']
            template_sheet[files_cell] = total_files
            template_sheet[usage_cell] = usage

    # Add LWX_Cost and LWN_Cost to the template
    template_sheet['F3'] = float(LWX_Cost)
    template_sheet['F3'].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    template_sheet['F4'] = float(LWN_Cost)
    template_sheet['F4'].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # Remove external links
    if hasattr(template_workbook, 'remove_external_links'):
        template_workbook.remove_external_links()

    #Write agency data to worksheets in the template
    for agency, rows in lwn_final_data.items():
        if not rows:
            continue
        sheet = template_workbook.create_sheet(title=agency)
        headers = list(rows[0].keys())  # Convert dict_keys to list
        sheet.append(headers)
        for row in rows:
            sheet.append([row[header] for header in headers])

    # Define the border style
    medium_border = Border(left=Side(style='medium'), 
                           right=Side(style='medium'), 
                           top=Side(style='medium'), 
                           bottom=Side(style='medium'),
                           outline=True)

    # Write calculated data to the template starting at row 32
    row = 32
    for department, totals in calculated_data.items():
        primary_total = totals.get('primary_total', 0)
        cloud_total = totals.get('cloud_total', 0)
        
        template_sheet[f'B{row}'] = department
        template_sheet[f'C{row}'] = primary_total
        template_sheet[f'D{row}'] = cloud_total
        template_sheet[f'E{row}'] = f'=C{row}+D{row}'
        # Define total_row before using it
        total_row = row + 1
        template_sheet[f'F{row}'] = f'=C{row}/C${total_row}'
        template_sheet[f'G{row}'] = f'=F{row}*$F$4'
        
        # Apply the border to the cells
        template_sheet[f'B{row}'].border = medium_border
        template_sheet[f'C{row}'].border = medium_border
        template_sheet[f'D{row}'].border = medium_border
        template_sheet[f'E{row}'].border = medium_border
        template_sheet[f'F{row}'].border = medium_border
        template_sheet[f'G{row}'].border = medium_border
        
        # Format column F as percentage with 2 decimal places
        template_sheet[f'F{row}'].number_format = '0.00%'
        # Format column G as currency with 2 decimal places
        template_sheet[f'G{row}'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        
        row += 1

    # Add a row for Totals
    total_row = row
    template_sheet[f'B{total_row}'] = 'Totals'
    template_sheet[f'C{total_row}'] = f'=SUM(C32:C{total_row-1})'
    template_sheet[f'D{total_row}'] = f'=SUM(D32:D{total_row-1})'
    template_sheet[f'E{total_row}'] = f'=C{total_row}+D{total_row}'
    template_sheet[f'F{total_row}'] = f'=C{total_row}/C{total_row}'
    template_sheet[f'G{total_row}'] = f'=F{total_row}*$F$4'
    
    # Update the percentage formula in column F to reference the total row
    for r in range(32, total_row):
        template_sheet[f'F{r}'] = f'=C{r}/C${total_row}'
        template_sheet[f'F{r}'].number_format = '0.00%'
        template_sheet[f'G{r}'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    
    # Apply the border to the Totals row
    template_sheet[f'B{total_row}'].border = medium_border
    template_sheet[f'C{total_row}'].border = medium_border
    template_sheet[f'D{total_row}'].border = medium_border
    template_sheet[f'E{total_row}'].border = medium_border
    template_sheet[f'F{total_row}'].border = medium_border
    template_sheet[f'G{total_row}'].border = medium_border

    # Save the updated workbook
    template_workbook.save(output_path)
    print(f"Data written to {output_path}")
